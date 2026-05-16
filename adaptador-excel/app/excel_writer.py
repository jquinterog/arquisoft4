import json
import logging
import os
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import Workbook, load_workbook


logger = logging.getLogger(__name__)

EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH", "data/promociones.xlsx")

HEADERS = [
    "event_id",
    "event_type",
    "source",
    "event_created_at",
    "promotion_id",
    "name",
    "description",
    "type",
    "status",
    "channel",
    "priority",
    "start_date",
    "end_date",
    "action_message",
    "raw_event",
]

_lock = Lock()


def ensure_workbook(path: str = EXCEL_FILE_PATH) -> None:
    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    if workbook_path.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "promociones"
    sheet.append(HEADERS)
    workbook.save(workbook_path)
    logger.info("excel_workbook_created", extra={"path": str(workbook_path)})


def append_event(event: dict[str, Any], path: str = EXCEL_FILE_PATH) -> dict[str, Any]:
    with _lock:
        ensure_workbook(path)
        workbook = load_workbook(path)
        sheet = workbook.active
        sheet.append(_event_to_row(event))
        row_count = sheet.max_row
        workbook.save(path)

    info = workbook_info(path)
    logger.info(
        "promotion_event_appended_to_excel event_id=%s path=%s data_rows=%s total_rows=%s",
        event.get("event_id"),
        info["path"],
        info["data_rows"],
        info["rows"],
    )
    return info


def workbook_info(path: str = EXCEL_FILE_PATH) -> dict[str, Any]:
    ensure_workbook(path)
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True)
    sheet = workbook.active
    return {
        "path": str(workbook_path),
        "sheet": sheet.title,
        "rows": sheet.max_row,
        "data_rows": max(sheet.max_row - 1, 0),
        "columns": sheet.max_column,
        "size_bytes": workbook_path.stat().st_size,
    }


def _event_to_row(event: dict[str, Any]) -> list[Any]:
    payload = event.get("payload") or {}
    return [
        event.get("event_id"),
        event.get("event_type"),
        event.get("source"),
        event.get("created_at"),
        payload.get("id"),
        payload.get("name"),
        payload.get("description"),
        payload.get("type"),
        payload.get("status"),
        payload.get("channel"),
        payload.get("priority"),
        payload.get("start_date"),
        payload.get("end_date"),
        payload.get("action_message"),
        json.dumps(event, ensure_ascii=True, default=str),
    ]
